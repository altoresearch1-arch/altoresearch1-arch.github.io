# -*- coding: utf-8 -*-
"""
fetch_bem.py — Producción minera mensual por empresa, del BEM del MINEM.

Fuente: "Cuadros Estadísticos Mineros – Prepublicación del BEM" (sección S01,
cuadro C02 "producción según empresa") y, para may/jun-2025 (sin prepublicación),
el "Excel de anexos" del Boletín (hoja "2. PRODUCCIÓN EMPRESAS").
Colección: https://www.gob.pe/institucion/minem/colecciones/6-boletin-estadistico-minero

Cómo funciona el dato: cada edición del mes M trae la producción de ese mes M
del año de la edición Y **y del mismo mes del año anterior (Y-1)**. Con las
ediciones jul-2025→abr-2026 (+anexos may/jun-2025) se arma la serie completa
ene-2025 → abr-2026. El cuadro por empresa es un TOP (~10) por metal + "OTROS":
si una empresa no aparece un mes, ese mes queda null (produjo poco o nada
relativo al top — la app lo dice con honestidad, Regla #1: no se inventa 0).

Salida: app/src/data/mineria.json
Caché: extractor/cache_bem/*.xlsx (no re-descarga lo que ya tiene).
Auto-descubre ediciones nuevas en la colección de gob.pe (para el robot).
"""
import io
import json
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path

import requests
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

BASE = Path(__file__).resolve().parent
CACHE = BASE / "cache_bem"
SALIDA = BASE.parent / "app" / "src" / "data" / "mineria.json"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                  "(KHTML, like Gecko) Chrome/126.0 Safari/537.36",
    "Accept-Language": "es-PE,es;q=0.9",
}

COLECCION_URL = "https://www.gob.pe/institucion/minem/colecciones/6-boletin-estadistico-minero"
GOB = "https://www.gob.pe"

# Ediciones conocidas: "YYYY-MM" de la edición -> URL del xlsx (s01 prepublicación
# o excel de anexos del boletín). Las nuevas se auto-descubren en la colección.
EDICIONES = {
    "2025-05": "https://cdn.www.gob.pe/uploads/document/file/8381452/6971828-bem-mayo-2025-excel-de-anexos.xlsx?v=1752789870",
    "2025-06": "https://cdn.www.gob.pe/uploads/document/file/8502940/7045145-bem-junio-2025-excel-de-anexos.xlsx?v=1755528922",
    "2025-07": "https://cdn.www.gob.pe/uploads/document/file/8622242/7135263-s01-produccion-minera-metalica-ed-07-2025%282%29.xlsx?v=1757600290",
    "2025-08": "https://cdn.www.gob.pe/uploads/document/file/8761202/7177188-s01-produccion-minera-metalica-ed-08-2025.xlsx?v=1759336112",
    "2025-09": "https://cdn.www.gob.pe/uploads/document/file/8987898/7393910-s01-produccion-minera-metalica-ed-09-2025.xlsx?v=1762983169",
    "2025-10": "https://cdn.www.gob.pe/uploads/document/file/9136707/7496853-s01-produccion-minera-metalica-ed-10-2025.xlsx?v=1765554626",
    "2025-11": "https://cdn.www.gob.pe/uploads/document/file/9295368/7571489-s01-produccion-minera-metalica-ed-11-2025.xlsx?v=1768511719",
    "2025-12": "https://cdn.www.gob.pe/uploads/document/file/9440997/7733058-s01-produccion-minera-metalica-ed-12-2025-vf.xlsx?v=1770935178",
    "2026-01": "https://cdn.www.gob.pe/uploads/document/file/9618397/7871419-s01-produccion-minera-metalica-ed-1-2026.xlsx?v=1773722791",
    "2026-02": "https://cdn.www.gob.pe/uploads/document/file/9763625/7955909-s01-produccion-minera-metalica-ed-02-2026.xlsx?v=1775660351",
    "2026-03": "https://cdn.www.gob.pe/uploads/document/file/9968538/8131158-s01-produccion-minera-metalica-ed-3-2026.xlsx?v=1778772864",
    "2026-04": "https://cdn.www.gob.pe/uploads/document/file/10070405/8203761-s01-produccion-minera-metalica-ed-4-2026.xlsx?v=1780408349",
}

# La serie que publica la app (se extiende sola cuando aparecen ediciones nuevas).
MES_INICIO = "2025-01"

MESES_ES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "setiembre": 9, "septiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}

METALES = {
    "COBRE": ("cobre", "TMF"),
    "ORO": ("oro", "g finos"),
    "ZINC": ("zinc", "TMF"),
    "PLATA": ("plata", "kg finos"),
    "PLOMO": ("plomo", "TMF"),
    "HIERRO": ("hierro", "TMF"),
    "ESTANO": ("estano", "TMF"),
    "MOLIBDENO": ("molibdeno", "TMF"),
    "ARSENICO": ("arsenico", "TMF"),
    "BISMUTO": ("bismuto", "TMF"),
    "CADMIO": ("cadmio", "TMF"),
    "MANGANESO": ("manganeso", "TMF"),
    "MAGNESIO": ("magnesio", "TMF"),
}

# Entidades del BEM que nos interesan (nombre normalizado -> clave).
# Cubre las empresas de la app + subsidiarias mineras de su familia.
ENTIDADES = {
    "NEXA RESOURCES PERU S.A.A.": "nexa_peru",
    "NEXA RESOURCES EL PORVENIR S.A.C.": "nexa_porvenir",
    "NEXA RESOURCES ATACOCHA S.A.A.": "atacocha",
    "COMPANIA DE MINAS BUENAVENTURA S.A.A.": "buenaventura",
    "MINERA LA ZANJA S.R.L.": "la_zanja",
    "COMPANIA MINERA COIMOLACHE S.A.": "coimolache",
    "SOCIEDAD MINERA CERRO VERDE S.A.A.": "cerro_verde",
    "SOCIEDAD MINERA EL BROCAL S.A.A.": "brocal",
    "VOLCAN COMPANIA MINERA S.A.A.": "volcan",
    "COMPANIA MINERA CHUNGAR S.A.C.": "chungar",
    "EMPRESA ADMINISTRADORA CERRO S.A.C.": "cerro_sac",
    "OXIDOS DE PASCO S.A.C.": "oxidos_pasco",
    "MINSUR S.A.": "minsur",
    "MARCOBRE S.A.C.": "marcobre",
    "SHOUGANG HIERRO PERU S.A.A.": "shougang",
    "MINERA SHOUXIN PERU S.A.": "shouxin",
    "SOUTHERN PERU COPPER CORPORATION SUCURSAL DEL PERU": "southern",
    "COMPANIA MINERA PODEROSA S.A.": "poderosa",
    "SOCIEDAD MINERA CORONA S.A.": "corona",
    "COMPANIA MINERA SANTA LUISA S.A.": "santa_luisa",
    "COMPANIA MINERA SAN IGNACIO DE MOROCOCHA S.A.A.": "simsa",
    "PERUBAR S.A.": "perubar",
}

NOMBRES_BONITOS = {
    "nexa_peru": "Nexa Resources Perú (Cerro Lindo)",
    "nexa_porvenir": "Nexa El Porvenir (subsidiaria de Nexa Perú)",
    "atacocha": "Nexa Atacocha",
    "buenaventura": "Buenaventura (minas propias)",
    "la_zanja": "La Zanja (subsidiaria de BVN)",
    "coimolache": "Coimolache / Tantahuatay (BVN 40.1%)",
    "cerro_verde": "Cerro Verde",
    "brocal": "El Brocal",
    "volcan": "Volcan (unidades propias)",
    "chungar": "Chungar (subsidiaria de Volcan)",
    "cerro_sac": "Adm. Cerro (subsidiaria de Volcan)",
    "oxidos_pasco": "Óxidos de Pasco (subsidiaria de Volcan)",
    "minsur": "Minsur (San Rafael / Pucamarca)",
    "marcobre": "Marcobre / Mina Justa (Minsur 60%)",
    "shougang": "Shougang Hierro Perú",
    "shouxin": "Minera Shouxin (relaves de Marcona; Shougang 49%)",
    "southern": "Southern Perú (sucursal)",
    "poderosa": "Poderosa",
    "corona": "Corona (Yauricocha)",
    "santa_luisa": "Santa Luisa",
    "simsa": "SIMSA (San Vicente)",
    "perubar": "Perubar",
}


def normalizar(texto):
    """MAYÚSCULAS sin tildes, espacios colapsados, sin asteriscos/notas."""
    t = unicodedata.normalize("NFD", str(texto))
    t = "".join(c for c in t if unicodedata.category(c) != "Mn")
    t = t.upper().replace("\n", " ")
    t = re.sub(r"\*+", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t


def descargar(mes, url):
    CACHE.mkdir(exist_ok=True)
    destino = CACHE / f"bem_{mes}.xlsx"
    if destino.exists() and destino.stat().st_size > 10000:
        return destino
    print(f"  bajando edición {mes}…")
    r = requests.get(url, headers=HEADERS, timeout=90)
    r.raise_for_status()
    destino.write_bytes(r.content)
    return destino


def parsear_edicion(ruta, mes_edicion):
    """Devuelve {(clave_entidad, metal): {mes: valor}} de una edición.

    La edición del mes M/año Y trae el valor de M/Y y de M/(Y-1).
    Funciona con el formato prepublicación (hoja S01.C02) y el de
    anexos del boletín (hoja '2. PRODUCCIÓN EMPRESAS').
    """
    anio, mes = mes_edicion.split("-")
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    hoja = None
    for nombre in wb.sheetnames:
        n = normalizar(nombre)
        if "S01.C02" in n or "PRODUCCION EMPRESAS" in n:
            hoja = wb[nombre]
            break
    if hoja is None:
        raise RuntimeError(f"{ruta.name}: no encuentro la hoja de producción por empresa")

    filas = [[c for c in fila] for fila in hoja.iter_rows(values_only=True)]
    wb.close()

    # Fila de cabecera: la que contiene 'PRODUCTO / EMPRESA'
    cidx = None
    inicio = None
    for i, fila in enumerate(filas):
        for j, celda in enumerate(fila):
            if celda and "PRODUCTO / EMPRESA" in normalizar(celda):
                cidx, inicio = j, i
                break
        if cidx is not None:
            break
    if cidx is None:
        raise RuntimeError(f"{ruta.name}: no encuentro la cabecera PRODUCTO / EMPRESA")

    # Columnas de años: las 2 siguientes celdas de esa fila con un año
    anios_cols = []
    for j in range(cidx + 1, min(cidx + 6, len(filas[inicio]))):
        celda = filas[inicio][j]
        if celda is not None and re.fullmatch(r"20\d\d", str(celda).strip()):
            anios_cols.append((int(str(celda).strip()), j))
        if len(anios_cols) == 2:
            break
    if len(anios_cols) != 2:
        raise RuntimeError(f"{ruta.name}: no encuentro las columnas de años")

    datos = {}
    metal_actual = None
    for fila in filas[inicio + 1:]:
        crudo = fila[cidx] if cidx < len(fila) else None
        if crudo is None or not str(crudo).strip():
            continue
        nombre = normalizar(crudo)
        if nombre.startswith("FUENTE"):
            break
        # ¿Es fila de metal? ej. 'COBRE (TMF) / COPPER (FTM)' o 'COBRE (TMF)'
        m = re.match(r"^([A-ZÑ ]+?)\s*\(", nombre)
        if m and m.group(1).strip() in METALES:
            metal_actual = METALES[m.group(1).strip()][0]
            # La fila del metal trae el TOTAL NACIONAL del mes → se guarda
            # bajo la clave especial '_pais' (la app calcula "% del Perú").
            for anio_col, j in anios_cols:
                valor = fila[j] if j < len(fila) else None
                if isinstance(valor, (int, float)):
                    mes_dato = f"{anio_col}-{mes}"
                    datos.setdefault(("_pais", metal_actual), {})[mes_dato] = round(float(valor), 2)
            continue
        if metal_actual is None or nombre.startswith("OTROS"):
            continue
        clave = ENTIDADES.get(nombre)
        if clave is None:
            continue
        for anio_col, j in anios_cols:
            valor = fila[j] if j < len(fila) else None
            if isinstance(valor, (int, float)):
                mes_dato = f"{anio_col}-{mes}"
                datos.setdefault((clave, metal_actual), {})[mes_dato] = round(float(valor), 2)
    return datos


def descubrir_nuevas():
    """Busca en la colección prepublicaciones más nuevas que las conocidas."""
    nuevas = {}
    try:
        r = requests.get(COLECCION_URL, headers=HEADERS, timeout=30)
        r.raise_for_status()
    except Exception as e:
        print(f"  (aviso: no pude leer la colección para buscar ediciones nuevas: {e})")
        return nuevas
    patron = re.compile(
        r'href="(/institucion/minem/informes-publicaciones/\d+-cuadros-estadisticos-'
        r'mineros-prepublicacion-del-bem-([a-z]+)-(\d{4}))"', re.I)
    for m in patron.finditer(r.text):
        ruta_pag, mes_nombre, anio = m.group(1), m.group(2).lower(), m.group(3)
        num = MESES_ES.get(mes_nombre)
        if not num:
            continue
        clave = f"{anio}-{num:02d}"
        if clave in EDICIONES or clave in nuevas:
            continue
        try:
            pag = requests.get(GOB + ruta_pag, headers=HEADERS, timeout=30)
            xls = re.findall(r'https://cdn\.www\.gob\.pe[^"]*s0?1[^"]*produccion[^"]*\.xlsx[^"]*', pag.text)
            if xls:
                nuevas[clave] = xls[0]
                print(f"  edición NUEVA descubierta: {clave}")
        except Exception as e:
            print(f"  (aviso: edición {clave} encontrada pero sin xlsx: {e})")
    return nuevas


def main():
    print("BEM MINEM — producción minera mensual por empresa")
    ediciones = dict(EDICIONES)
    ediciones.update(descubrir_nuevas())

    acumulado = {}   # (clave, metal) -> {mes: valor}
    ok = 0
    for mes_ed in sorted(ediciones):
        try:
            ruta = descargar(mes_ed, ediciones[mes_ed])
            datos = parsear_edicion(ruta, mes_ed)
            for k, meses in datos.items():
                acumulado.setdefault(k, {}).update(meses)
            ok += 1
        except Exception as e:
            print(f"  ❌ edición {mes_ed}: {e}")
    print(f"  {ok}/{len(ediciones)} ediciones procesadas")

    # Eje de meses: MES_INICIO → última edición
    ultima = max(ediciones)
    meses = []
    a, m = map(int, MES_INICIO.split("-"))
    while f"{a}-{m:02d}" <= ultima:
        meses.append(f"{a}-{m:02d}")
        m += 1
        if m > 12:
            m, a = 1, a + 1

    entidades = {}
    totales_pais = {}
    for (clave, metal), pormes in sorted(acumulado.items()):
        if clave == "_pais":
            totales_pais[metal] = [pormes.get(mes) for mes in meses]
            continue
        ent = entidades.setdefault(clave, {
            "nombre": NOMBRES_BONITOS.get(clave, clave),
            "produccion": {},
        })
        ent["produccion"][metal] = [pormes.get(mes) for mes in meses]

    salida = {
        "actualizado": date.today().isoformat(),
        "fuente": "MINEM — Boletín Estadístico Minero (producción por empresa, top de productores por metal)",
        "fuenteUrl": COLECCION_URL,
        "nota": ("Si un mes no hay dato, la empresa no apareció ese mes entre los "
                 "principales productores del metal (produjo poco o nada) — no se inventa un cero."),
        "meses": meses,
        "unidades": {v[0]: v[1] for v in METALES.values()},
        "totalesPais": totales_pais,
        "entidades": entidades,
    }
    # El BEM cambia UNA vez al mes: si los datos no cambiaron, no se reescribe
    # (evita que el robot nocturno commitee un archivo idéntico cada día).
    if SALIDA.exists():
        try:
            previo = json.loads(SALIDA.read_text(encoding="utf-8"))
            if {k: v for k, v in previo.items() if k != "actualizado"} == \
               {k: v for k, v in salida.items() if k != "actualizado"}:
                print(f"  ✅ {SALIDA.name}: sin cambios (el BEM no publicó nada nuevo)")
                return
        except Exception:
            pass
    SALIDA.write_text(json.dumps(salida, ensure_ascii=False, indent=1), encoding="utf-8")

    n_series = sum(len(e["produccion"]) for e in entidades.values())
    print(f"  ✅ {SALIDA.name}: {len(entidades)} entidades, {n_series} series, meses {meses[0]} → {meses[-1]}")
    # Resumen por entidad
    for clave, ent in sorted(entidades.items()):
        mets = ", ".join(f"{m}({sum(1 for x in ent['produccion'][m] if x is not None)})"
                         for m in ent["produccion"])
        print(f"    {clave}: {mets}")


if __name__ == "__main__":
    main()
