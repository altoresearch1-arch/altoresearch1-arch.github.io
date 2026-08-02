# -*- coding: utf-8 -*-
"""📄 Leer el texto de un documento de la SMV, sea del formato que sea.

No todas las empresas presentan en PDF. Al 01-ago-2026, de las 90 que ya tenían
el Q2, DIEZ subieron su "Análisis y Discusión de la Gerencia" en Word o Excel
(siete en .doc/.xls antiguo y tres en .docx/.xlsx), y varias hicieron lo mismo
con las Notas a los EE.FF. Antes reventaban con "Stream has ended unexpectedly"
—que parece un corte de red y no lo es— y esas fichas se quedaban con el texto
del trimestre anterior al lado de cifras nuevas, contradiciéndose solas.

Vive en su propio módulo (y no dentro de fetch_gerencia) porque fetch_notas
también lo necesita: importar un script que al cargarse reconfigura sys.stdout
cierra el buffer del que ya lo hizo, y el segundo proceso muere con
"I/O operation on closed file".
"""
import io, re, zipfile

from pypdf import PdfReader


def leer_documento(raw):
    """Devuelve (texto | None, paginas, formato). texto None = formato desconocido."""
    if raw[:4] == b"%PDF":
        lector = PdfReader(io.BytesIO(raw))
        return ("\n".join((p.extract_text() or "") for p in lector.pages),
                len(lector.pages), "PDF")

    if raw[:2] == b"PK":                       # .docx / .xlsx (Office moderno = ZIP)
        with zipfile.ZipFile(io.BytesIO(raw)) as z:
            nombres = z.namelist()
            if "word/document.xml" in nombres:
                xml = z.read("word/document.xml").decode("utf-8", "ignore")
                # cada </w:p> es un párrafo; el resto de etiquetas se cae
                xml = re.sub(r"</w:p>", "\n", xml)
                return re.sub(r"<[^>]+>", "", xml), 1, "DOCX"
            if any(n.startswith("xl/") for n in nombres):
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                filas = []
                for hoja in wb.worksheets:
                    for fila in hoja.iter_rows(values_only=True):
                        celdas = [str(v) for v in fila if v is not None]
                        if celdas:
                            filas.append(" ".join(celdas))
                return "\n".join(filas), len(wb.worksheets), "XLSX"
        return None, 0, "ZIP desconocido"

    if raw[:4] == b"\xd0\xcf\x11\xe0":         # .doc / .xls antiguo (OLE2)
        import olefile
        o = olefile.OleFileIO(io.BytesIO(raw))
        streams = {"/".join(x) for x in o.listdir()}
        if "WordDocument" in streams:
            # El .doc guarda el texto entre estructuras binarias. Se queda solo con
            # lo imprimible; la basura del encabezado no sobrevive al filtro de
            # frases, que exige cifras o palabras clave para conservar una oración.
            crudo = o.openstream("WordDocument").read().decode("latin-1", "ignore")
            limpio = re.sub(r"[^\x20-\x7E\xC0-\xFF\n]+", " ", crudo)
            return re.sub(r" {2,}", " ", limpio), 1, "DOC"
        if "Workbook" in streams or "Book" in streams:
            import xlrd
            libro = xlrd.open_workbook(file_contents=raw)
            filas = []
            for hoja in libro.sheets():
                for i in range(hoja.nrows):
                    celdas = [str(v) for v in hoja.row_values(i) if v not in ("", None)]
                    if celdas:
                        filas.append(" ".join(celdas))
            return "\n".join(filas), libro.nsheets, "XLS"
        return None, 0, "OLE2 desconocido"

    return None, 0, f"desconocido {raw[:4]!r}"
